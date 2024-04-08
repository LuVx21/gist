# coding=utf-8

# openpyxl可以对excel文件进行读写操作

from openpyxl import load_workbook
from openpyxl import Workbook
# from openpyxl.writer.excel import ExcelWriter

workbook = load_workbook(u"1.xlsx")
sheetnames = workbook.get_sheet_names()
print(sheetnames)

sheet = workbook.get_sheet_by_name(sheetnames[0])
# 行列从1开始
print(sheet.cell(row = 1 , column = 1).value)
sheet['A1'] = '47'
workbook.save(u"2.xlsx")

workbookNew = Workbook()
ws = workbookNew.active
ws['A1'] = 47
workbookNew.save("3.xlsx")