#!/usr/bin/python3
# -*- coding:utf-8 -*-
import pymysql
import xlwt
from openpyxl import load_workbook
from openpyxl import Workbook


# 打开数据库连接
connection = pymysql.connect("localhost", "root", "1121", "python")

# 使用 cursor() 方法创建一个游标对象 cursor
cursor = connection.cursor()

# SQL 查询语句
sql = "select * from python order by lum1"

try:
    # 执行SQL语句
    temp = cursor.execute(sql)
    # 获取所有记录列表
    rows = cursor.fetchall()
    print("获取到%d条记录." % temp)
except:
    print("Error: unable to fecth data")

# 关闭数据库连接
connection.close()

filePath = u"/Users/renxie/Code/Python/Python36/src/MySQL/1.xls"
filePath1 = u"/Users/renxie/Code/Python/Python36/src/MySQL/1.xlsx"


def writer(colnum=0, rownum=0, rows=None, sheet=None):
    """
    查询结果写入新建excel中
    :param colnum: 起点列
    :param rownum: 起点行
    :param rows: 查询结果
    :param sheet: 写入excel对象的sheet
    :return:
    """
    colnumtmp = colnum
    for row in rows:
        for item in row:
            sheet.write(rownum, colnum, item)
            colnum = colnum + 1
        colnum = colnumtmp
        rownum = rownum + 1


# xlwt方式写入
workbook = xlwt.Workbook()
sheet = workbook.add_sheet('sheet1')
writer(1, 1, rows, sheet)
workbook.save(filePath)


def writer1(colnum=0, rownum=0, rows=None, sheet=None):
    """
    查询结果写入给定excel中
    :param colnum: 起点列
    :param rownum: 起点行
    :param rows: 查询结果
    :param sheet: 写入excel对象的sheet
    :return:
    """
    colnum = colnum + 64
    colnumtmp = colnum
    for row in rows:
        for item in row:
            location = chr(colnum) + str(rownum)
            sheet[location] = item
            colnum = colnum + 1
        colnum = colnumtmp
        rownum = rownum + 1


# openpyxl方式写入
workbookNew = Workbook()
ws = workbookNew.active
writer1(1, 1, rows, ws)
workbookNew.save(filePath1)